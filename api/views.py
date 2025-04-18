from rest_framework import generics
import logging
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import Event, Currency, User, UserCurrency, UserCurrencySummary
from .serializers import EventSerializer, CurrencySerializer, UserSerializer, UserCurrencySerializer
from django.shortcuts import render
logger = logging.getLogger(__name__)
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework import status
import logging
logger = logging.getLogger(__name__)
from rest_framework_simplejwt.views import TokenRefreshView
from rest_framework_simplejwt.serializers import TokenRefreshSerializer

class EventList(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Event.objects.all()
    serializer_class = EventSerializer

    def get(self, request, *args, **kwargs):
        logger.debug("Received GET request")
        return super().get(request, *args, **kwargs)
    
    def put(self, request, *args, **kwargs):
        event_id = kwargs.get('pk')
        try:
            event = Event.objects.get(pk=event_id)
            event.type = request.data.get('type')
            event.currency = request.data.get('currency')
            event.amount = request.data.get('amount')
            event.rate = request.data.get('rate')
            event.total = request.data.get('total')
            event.save()
            return Response(status=status.HTTP_200_OK)
        except Event.DoesNotExist:
            logger.error(f"Event with ID {event_id} not found")
            return Response(
                {"error": "Event not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error updating event: {str(e)}")
            return Response(
                {"error": "Failed to update event"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request, *args, **kwargs):
        logger.debug(f"Received POST data: {request.data}")
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.error(f"Validation errors: {serializer.errors}")
        return Response(
            {
                "error": "Invalid data",
                "details": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, *args, **kwargs):
        event_id = kwargs.get('pk')
        logger.debug(f"Attempting to delete event with ID: {event_id}")

        try:
            event = Event.objects.get(pk=event_id)
            event.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Event.DoesNotExist:
            logger.error(f"Event with ID {event_id} not found")
            return Response(
                {"error": "Event not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error deleting event: {str(e)}")
            return Response(
                {"error": "Failed to delete event"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class CurrencyList(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CurrencySerializer

    def get_queryset(self):
        if self.request.method == 'GET':
            return Currency.objects.values('name', 'rate_to_som')
        return Currency.objects.all()

    def post(self, request, *args, **kwargs):
        currency_name = request.data.get('name')
        rate_to_som = request.data.get('rate_to_som')

        if not currency_name:
            return Response(
                {"error": "Currency name is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if Currency.objects.filter(name=currency_name).exists():
                return Response(
                    {"error": "Currency already exists"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            currency = Currency.objects.create(
                name=currency_name,
                rate_to_som=rate_to_som
            )
            serializer = self.get_serializer(currency)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error(f"Error creating currency: {str(e)}")
            return Response(
                {"error": "Failed to create currency"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def delete(self, request, *args, **kwargs):
        currency_name = request.data.get('name')
        try:
            currency = Currency.objects.get(name=currency_name)
            currency.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Currency.DoesNotExist:
            logger.error(f"Currency with name {currency_name} not found")
            return Response(
                {"error": "Currency not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error deleting currency: {str(e)}")
            return Response(
                {"error": "Failed to delete currency"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def put(self, request, *args, **kwargs):
        currency_name = request.data.get('newName')
        currency_old_name = request.data.get('oldName')
        new_rate = request.data.get('rate_to_som')  # опционально обновляем курс

        try:
            currency = Currency.objects.get(name=currency_old_name)
            currency.name = currency_name
            if new_rate is not None:
                currency.rate_to_som = new_rate
            currency.save()
            return Response(status=status.HTTP_200_OK)
        except Currency.DoesNotExist:
            logger.error(f"Currency with name {currency_old_name} not found")
            return Response(
                {"error": "Currency not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error updating currency: {str(e)}")
            return Response(
                {"error": "Failed to update currency"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class UsersList(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get(self, request, *args, **kwargs):
        if not kwargs.get('username'):
            return super().get(request, *args, **kwargs)
        user_id = kwargs.get('username')
        return Response(
            UserSerializer(User.objects.get(username=user_id)).data,
            status=status.HTTP_200_OK
        )

    def post(self, request, *args, **kwargs):
        logger.debug(f"Received POST data: {request.data}")
        username = request.data.get('username')
        password = request.data.get('password')
        
        # Check if user already exists
        if User.objects.filter(username=username).exists():
            return Response(
                {
                    "error": "Имя пользователя уже занято",
                    "details": "A user with this username already exists"
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        elif User.objects.filter(email=request.data.get('email')).exists():
            return Response(
                {
                    "error": "Эта почта уже зарегистрирована",
                    "details": "A user with this email already exists"
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        elif User.objects.filter(phone=request.data.get('phone')).exists():
            return Response(
                {
                    "error": "Этот номер уже зарегистрирован",
                    "details": "A user with this phone number already exists"
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
            
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            # Get validated data
            validated_data = serializer.validated_data
            # Remove password from validated data
            password = validated_data.pop('password', None)
            # Create user instance
            user = User(**validated_data)
            # Set password properly to ensure it's hashed
            user.set_password(password)
            user.save()
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        logger.error(f"Validation errors: {serializer.errors}")
        return Response(
            {
                "error": "Invalid data",
                "details": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    def put(self, request, *args, **kwargs):
        new_user_id = request.data.get('username')
        old_user_id = request.data.get('oldUsername')
        is_superuser = request.data.get('isSuperUser')
        email = request.data.get('email')
        logger.debug(is_superuser)
        print(is_superuser)
        if request.data.get('password'):
            new_password = request.data.get('password')
        else:
            new_password = None
        try:
            user = User.objects.get(username=old_user_id)
            if new_password:
                user.set_password(new_password)
            user.username = new_user_id
            user.is_superuser = True if is_superuser else False
            user.email = email
            user.save()
            return Response(status=status.HTTP_200_OK)
        except User.DoesNotExist:
            logger.error(f"User with ID {old_user_id} not found")
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error updating user: {str(e)}")
            return Response(
                {"error": "Failed to update user"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def delete(self, request, *args, **kwargs):
        user_id = request.data.get('username')
        logger.debug(f"Attempting to delete user with ID: {user_id}")
        try:
            user = User.objects.get(username=user_id)
            user.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except User.DoesNotExist:
            logger.error(f"User with ID {user_id} not found")
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error deleting user: {str(e)}")
            return Response(
                {"error": "Failed to delete user"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class UserCurrencyList(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserCurrencySerializer

    def get_queryset(self):
        return UserCurrency.objects.filter(user=self.request.user)

    def post(self, request, *args, **kwargs):
        name = request.data.get('name')
        rate_to_som = request.data.get('rate_to_som')
        amount = request.data.get('amount')
        event_date = request.data.get('event_date')
        event_type = request.data.get('event_type')

        if not name or not rate_to_som or not amount or not event_date or not event_type:
            return Response({"error": "Не указано одно из обязательных полей (name, rate_to_som, amount, event_date, event_type)"},
                            status=status.HTTP_400_BAD_REQUEST)

        if event_type not in ['purchase', 'sale']:
            return Response({"error": "Неверный тип события. Должно быть 'purchase' или 'sale'"},
                            status=status.HTTP_400_BAD_REQUEST)

        currency, created = Currency.objects.get_or_create(name=name)

        if created:
            currency.rate_to_som = rate_to_som
            currency.save()

        existing_currency = UserCurrency.objects.filter(user=request.user, currency=currency).first()
        if existing_currency:
            return Response({"error": "Вы уже задали курс для этой валюты"},
                            status=status.HTTP_400_BAD_REQUEST)

        if event_type == 'purchase':
            purchase_total = amount * rate_to_som
            purchase_count = amount
            sale_total = 0
            sale_count = 0
        elif event_type == 'sale':
            sale_total = amount * rate_to_som
            sale_count = amount
            purchase_total = 0
            purchase_count = 0

        purchase_average = purchase_total / purchase_count if purchase_count > 0 else 0
        sale_average = sale_total / sale_count if sale_count > 0 else 0
        profit = sale_total - purchase_total

        user_currency = UserCurrency(
            user=request.user,
            currency=currency,
            rate=rate_to_som,
            event_date=event_date,
            event_type=event_type,
            amount=amount,
            purchase_total=purchase_total,
            purchase_count=purchase_count,
            sale_total=sale_total,
            sale_count=sale_count
        )
        user_currency.save()
        print(request.data)

        return Response(UserCurrencySerializer(user_currency).data, status=status.HTTP_201_CREATED)

class UserCurrencyDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserCurrencySerializer

    def get_queryset(self):
        return UserCurrency.objects.filter(user=self.request.user)

    def put(self, request, *args, **kwargs):
        currency_id = kwargs.get('pk')
        rate = request.data.get('rate')
        amount = request.data.get('amount')
        event_type = request.data.get('event_type')

        if not rate or not amount or not event_type:
            return Response({"error": "Не указаны все обязательные поля"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user_currency = UserCurrency.objects.get(id=currency_id, user=request.user)
            user_currency.rate = rate
            user_currency.amount = amount
            user_currency.event_type = event_type
            user_currency.save()

            if event_type == 'purchase':
                user_currency.update_totals_and_profit(purchase_amount=amount)
            elif event_type == 'sale':
                user_currency.update_totals_and_profit(sale_amount=amount)

            return Response(UserCurrencySerializer(user_currency).data, status=status.HTTP_200_OK)
        except UserCurrency.DoesNotExist:
            return Response({"error": "Курс валюты не найден"}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, *args, **kwargs):
        currency_id = kwargs.get('pk')

        try:
            user_currency = UserCurrency.objects.get(id=currency_id, user=request.user)
            user_currency.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except UserCurrency.DoesNotExist:
            return Response({"error": "Курс валюты не найден"}, status=status.HTTP_404_NOT_FOUND)

class UserOwnedCurrenciesView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CurrencySerializer

    def get_queryset(self):
        return Currency.objects.filter(user_rates__user=self.request.user).distinct()

class ClearAll(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        try:
            user_currencies = UserCurrency.objects.filter(user=request.user)
            user_currencies.delete()

            Currency.objects.all().delete()

            Event.objects.all().delete()

            return Response({"message": "Все данные о валюте успешно очищены."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ExportUserCurrenciesView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user

        if not user.is_authenticated:
            return HttpResponse('Неавторизован', status=401)

        wb = Workbook()
        ws = wb.active
        big_font = Font(size=14)

        headers = [
            'Имя пользователя', 'Валюта', 'Курс', 'Дата события', 'Тип события', 'Количество',
            'Сумма покупки', 'Количество покупок', 'Сумма продажи', 'Количество продаж'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = big_font

        user_currencies = UserCurrency.objects.filter(user=user)

        for user_currency in user_currencies:
            row = [
                user.username,
                user_currency.currency.name,
                user_currency.rate,
                user_currency.event_date.strftime('%Y-%m-%d %H:%M:%S') if user_currency.event_date else '',
                user_currency.event_type,
                user_currency.amount,
                user_currency.purchase_total,
                user_currency.purchase_count,
                user_currency.sale_total,
                user_currency.sale_count
            ]
            ws.append(row)

            for cell in ws[ws.max_row]:
                cell.font = big_font

        # Устанавливаем минимальную ширину вручную
        column_widths = {
            'A': 22,
            'B': 15,
            'C': 10,
            'D': 25,
            'E': 18,
            'F': 14,
            'G': 18,
            'H': 23,
            'I': 18,
            'J': 22,
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="user_currencies.xlsx"'
        wb.save(response)
        return response

class EventListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EventSerializer

    def get_queryset(self):
        # Фильтруем события только для текущего пользователя
        return Event.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        # Сохраняем объект с текущим пользователем
        serializer.save(user=self.request.user)

    def post(self, request, *args, **kwargs):
        # Здесь обработка POST-запроса для создания нового события
        serializer = EventSerializer(data=request.data)
        if serializer.is_valid():
            self.perform_create(serializer)  # Сохраняем с текущим пользователем
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class EventDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EventSerializer

    def get_queryset(self):
        return Event.objects.all()

    def put(self, request, *args, **kwargs):
        event = self.get_object()
        serializer = EventSerializer(event, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, *args, **kwargs):
        event = self.get_object()
        event.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class ExportEventsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.append(['Тип', 'Валюта', 'Сумма', 'Дата', 'Время', 'Курс', 'Итого'])

        events = Event.objects.all()

        for event in events:
            ws.append([
                event.type,
                event.currency,
                float(event.amount),
                event.date,
                event.time.strftime("%H:%M:%S"),
                float(event.rate),
                float(event.total)
            ])

        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 20  # Тип
        ws.column_dimensions['B'].width = 15  # Валюта
        ws.column_dimensions['C'].width = 15  # Сумма
        ws.column_dimensions['D'].width = 15  # Дата
        ws.column_dimensions['E'].width = 15  # Время
        ws.column_dimensions['F'].width = 10  # Курс
        ws.column_dimensions['G'].width = 15  # Итого

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="events.xlsx"'
        wb.save(response)
        return response

class UserCurrencySummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Получаем все события пользователя
        user_events = Event.objects.filter(user=request.user)
        print(f"Found {user_events.count()} events for user {request.user}")

        # Получаем уникальные валюты
        currencies = user_events.values_list('currency', flat=True).distinct()
        print(f"Found currencies: {currencies}")

        summary_list = []
        for currency_name in currencies:
            print(f"Processing currency: {currency_name}")
            summary = UserCurrencySummary(request.user, currency_name)
            summary_list.append(summary.as_dict())

        print(f"Returning summary: {summary_list}")
        return Response(summary_list)

class isSuperAdmin(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        username = kwargs.get('username')
        try:
            superAdmin = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response({"error": "Not superadmin"}, status=status.HTTP_400_BAD_REQUEST)
        if superAdmin.is_superuser:
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({"error": "Not superadmin"}, status=status.HTTP_400_BAD_REQUEST)

class testRenderResetTemplateUi(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        return render(request, 'password_reset_confirm.html', {'validlink': True, 'uidb64': 'uidb64', 'token': 'token'})

class CustomTokenRefreshView(TokenRefreshView):
    serializer_class = TokenRefreshSerializer

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)

        # Можешь логировать или добавлять свои поля в ответ:
        response.data['message'] = 'Токен успешно обновлён'
        return response